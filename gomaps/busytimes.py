# -*- coding: utf-8 -*-
'''
Author: David J. Morfe
Application Name: busytimes
Functionality Purpose: Web scrape Google Maps Popular Times data
Version: Beta
'''
#8/7/20

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook, load_workbook
from datetime import datetime
from urllib.request import urlopen, urlparse, urlretrieve
from urllib.parse import quote_plus, unquote_plus
from .gmapss import GoogleMaps
import GeoLiberator as GL
import re
import pandas as pd
import os
import time
try:
    from bs4 import BeautifulSoup
except ImportError:
    from BeautifulSoup import BeautifulSoup

#Populate each day with 24 hour-clock
#Append date timestamps for each week
    #ex: 1/1/2019 8:30:00
#Modularize & Optimize (lighten weight by getting rid of selenium and openpyxl)

#Account for multiple map results
#Append Coordinates, City, State
#Speed optimization
#Web scrape maximum fire dept building occupancy capacity

exceptions = ["""
    Your search - {} - did not match any Google Maps Sites.
    Suggestions:
    • Addresses don't work, only place names.
    • Make sure all words are spelled correctly.
    • Try adding a city, state, or zip code to the place name."""]

class LocationNotFoundError(Exception):
    pass

class BusyTimes:
    '''
    A class for instantiating a BusyTimes object.
    The BusyTimes object takes in a location as input and outputs popular times data for that week.
    The popular times are based on average popularity over the past several weeks.
    The percentages are the peak times for that hour.
    '''

    def __init__(self, location):
        self.location = str(location)

    def __get_driver(self, driver_path: str):
        if "phantomjs" in driver_path:
            driver = webdriver.PhantomJS(executable_path=driver_path)
        else:
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            driver = webdriver.Chrome(executable_path="chromedriver.exe", chrome_options=chrome_options)
        return driver
    def _put_time(self, val):
        current_hour = re.search(r"^\d+", str(datetime.now().time())).group()
        if int(current_hour) < 12:
            current_hour += " AM"
        else:
            current_hour += " PM"
        if "Currently" in val:
            return val.strip('.') + f" at {current_hour}"
        else:
            return val

    def _get_missing_times(self, lst: list):
        def incrementTime(num, apm, bol=True):
            ret = 0
            if bol == True:
                if num == 11 and apm == "AM":
                    apm = "PM"
                elif num == 11 and apm == "PM":
                    apm = "AM"
                ret = num + 1
                if ret == 13:
                    ret = 1
            else:
                if num == 12 and apm == "PM":
                    apm = "AM"
                elif num == 12 and apm == "AM":
                    apm = "PM"
                ret = num - 1
                if ret == 0:
                    ret = 12
            return f"{ret} {apm}"
        ret = 0; pos = 0; 
        if not re.search(r"\d+ [AP]M", str(lst[0])):
            got = re.search(r"(\d+) ([AP]M)", str(lst[1]))
            ret = incrementTime(int(got.group(1)), got.group(2), False)
        elif not re.search(r"\d+ [AP]M", str(lst[-1])):
            got = re.search(r"(\d+) ([AP]M)", str(lst[-2]))
            ret = incrementTime(int(got.group(1)), got.group(2)); pos = -1
        else:
            for i in lst:
                get = re.search(r"(\d+) ([AP]M)", str(i))
                if get:
                    ret = (int(get.group(1)), get.group(2))
                else:
                    ret = incrementTime(ret[0], ret[1]); pos = lst.index(i)
                    break
        return (ret, pos)

    def _scrape_times(self, driver_path: str, keep_driver_alive: bool=False) -> dict:
        driver = self.__get_driver(driver_path)
        driver.implicitly_wait(30)
        check_link = re.search(r"www.google.com/maps(/|\?q=)", self.location)
        data = {}; c = 0
        days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        if check_link: #Check if location is already link
            driver.get(self.location); curl = driver.current_url
            if check_link.group(1) == "?q=":
                while True:
                    if curl != driver.current_url:
                        driver.get(driver.current_url)
                        driver.implicitly_wait(30)
                        break
        else: #Get location's google maps link
            url = get_query_link(self.location)
            if url == '':
                driver.quit()
                raise LocationNotFoundError(exceptions[0].format(self.location))
            nurl = GoogleMaps(url).url
            driver.get(nurl)
        soup = BeautifulSoup(driver.page_source, 'lxml')
        for item in soup.find_all("div", "section-bad-query-title"):
            if re.search(r"Maps can't find", str(item)):
                driver.quit()
                raise LocationNotFoundError(exceptions[0].format(self.location))
        popular_times_elem = driver.find_element_by_class_name("section-popular-times-container")
        day_of_week_elem = popular_times_elem.find_elements_by_class_name("section-popular-times-graph")
        for day in day_of_week_elem:
            day_items = day.find_elements_by_class_name("section-popular-times-bar")
            hour_items = [self._put_time(i.get_attribute("aria-label")) for i in day_items if i.get_attribute("aria-label") != None]
            data[days[c]] = hour_items
            c += 1
        if not keep_driver_alive:
            driver.quit()
        return data

def _convert_to_24_hour(times):
    return times #return new list as 24 hour time format

def _populate24Hours(times):
    all_hours = [f"0% busy at {i}" for i in range(24)]
    for i, v in enumerate(times):
        if v == "% busy at .":
            times = all_hours
        else:
            times[i] = times[i]
    return times #return new list with 24 hours per day AM/PM

def popular_times(location: str, driver, file: str='', keep_driver_alive: bool=False) -> dict:
    '''This function searches for a location or place on Google Maps, and returns its popular times data.

    :param location: The name of a place used to search on Google Maps
    :param driver: The path to the web driver being utilized
    :param file: The file to write the popular times data to
    :param keep_driver_alive: If True, keeps driver application open so it doesn't have to close and reopen again in loop.
    :type location: str
    :type driver: str
    :type file: str
    :type keep_driver_alive: bool

    :returns: A dictionary of the popular times data with the days of the week as keys and the values being a list of times with the percentage of how busy that place is at said time.
    '''
    if keep_driver_alive:
        print("Warning: Web Driver terminal window will have to be closed manually!")
    result = BusyTimes(location)._scrape_times(driver, keep_driver_alive)
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    for day in result:
        result[day] = _populate24Hours(result[day])
    if file == '': #Return python dictionary if not writing to file
        return result
    else: #Write to file
        pat = re.search(r"https://www.google.com/maps(/place)?", location)
        if pat: #Retrieve query name from link
            location = urlparse(location)
            if pat.group(1) == "/place":
                location = re.sub(r"/.+", '', unquote_plus(location[2]).strip("/maps/place/"))
            else:
                if re.search(r"\?.+?=", location.geturl()):
                    location = re.sub(r"api=1&query=|\?q=|&.*", '', unquote_plus(location[4])).strip(', us')
        if not os.path.isfile(file):
            df = pd.DataFrame(columns=["Place Name", "Day of Week", "Hour of Day", "Popular Time Value"])
            df.to_excel(file, index=False)
        df = pd.DataFrame(columns=["Place Name", "Day of Week", "Hour of Day", "Popular Time Value"])
        for k in result: #Populate DataFrame
            for i in result[k]:
                val = ''; tim = ''
                if i != "% busy at .":
                    val = re.search(r"^\d+%|(?<=Currently )\d+%", i).group()
                    tim = re.sub(r" ", ":00:00 ", (re.search(r"\d+ [AP]M", i).group()))
                    if "Currently" in i:
                        tim = tim + " (LIVE)"
                else:
                    val = "Closed"
                    tim = "N/A"
                df = df.append({"Place Name": location, "Day of Week": k, "Hour of Day": tim, "Popular Time Value": val}, ignore_index=True)
        if ".xls" in file:
            append_df_to_xl(file, df)
        elif ".csv" in file:
            print("W.I.P.")
        elif ".txt" in file:
            print("W.I.P.")
        else:
            raise Exception("File type not supported")

def get_query_link(place: str) -> str:
    '''This function returns a new link to query Google Maps

    :param place: Name of a place to be queried on Google Maps
    :type place: str

    :returns: The Google Maps URL to be send a request to
    '''
    link = '';
    if GL.parse_address(place, "address") == "OTHER":
        place = quote_plus(place)
        link = "https://www.google.com/maps?q=" + place
    return link

def append_df_to_xl(file_name: str, data):
    '''This function appends data to an Excel file

    :param file_name: Excel file name to write to
    :param data: Python Pandas DataFrame
    :type file_name: str
    :type data: pandas.DataFrame
    '''
    writer = pd.ExcelWriter(file_name, engine="openpyxl", mode='a')
    book = load_workbook(file_name)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    row_cursor = writer.book["Sheet1"].max_row
    data.to_excel(writer, "Sheet1", index=False, header=False, startrow=row_cursor)
    writer.save()
